# app.py - VERSIÓN SIMPLE Y FUNCIONANDO
from __future__ import annotations
from pathlib import Path
from typing import Any
import pandas as pd
from flask import Flask, jsonify, render_template, request, session, send_file
from functools import wraps
import hashlib
from datetime import datetime
import io
from models import ReviewDatabase
import json
import re

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

db = ReviewDatabase()

SUPERVISORS = {
    14: "Bruno Del Popolo",
    17: "Franco Vivani",
    41: "Claudio Raposo",
}

app = Flask(__name__)
app.secret_key = "clave_secreta_para_desarrollo"

_cached_df = None
_current_excel = None
_data_loaded = False
_current_mes = None

# =====================================================
# CARGAR CLIENTES DESDE JSON (FUNCIONA SIEMPRE)
# =====================================================

def cargar_clientes_json():
    """Cargar clientes desde clientes.json"""
    try:
        json_path = BASE_DIR / "clientes.json"
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                clientes = json.load(f)
            print(f"✅ Clientes cargados desde JSON: {len(clientes)}")
            return clientes
    except Exception as e:
        print(f"❌ Error cargando JSON: {e}")
    return {}

CLIENTES = cargar_clientes_json()

# =====================================================
# FUNCIONES
# =====================================================

def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()

def _is_visita_valida(value: Any) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    if isinstance(value, bool):
        return value
    text = str(value).strip().upper()
    return text in {"VERDADERO", "TRUE", "1", "1.0", "SI", "YES"}

def extract_short_poc_id(poc_id):
    """Extraer código corto del POC ID"""
    if not poc_id:
        return ""
    
    poc_id = str(poc_id).strip()
    if poc_id.endswith('.0'):
        poc_id = poc_id[:-2]
    poc_id = poc_id.replace('-', '').replace(' ', '')
    
    # Prefijo 053822 (con 0) o 53822 (sin 0)
    if poc_id.startswith("053822"):
        resultado = poc_id[6:].lstrip('0')
        return resultado if resultado else "0"
    if poc_id.startswith("53822"):
        resultado = poc_id[5:].lstrip('0')
        return resultado if resultado else "0"
    
    return poc_id.lstrip('0') or "0"

def get_client_info(poc_id):
    """Obtener datos del cliente desde JSON"""
    if not poc_id:
        return None
    
    short_id = extract_short_poc_id(poc_id)
    if not short_id or short_id == "0":
        return None
    
    if short_id in CLIENTES:
        return CLIENTES[short_id]
    
    return None

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'supervisor_id' not in session:
            return jsonify({"error": "No autorizado"}), 401
        return f(*args, **kwargs)
    return decorated_function

def _build_task_id(row: pd.Series) -> str:
    """Crear ID único para cada tarea usando POC ID + Fecha + Detalle Tarea"""
    poc_id = clean_text(row.get("POC ID", ""))
    fecha = str(row.get("Fecha", ""))
    detalle = clean_text(row.get("Detalle Tarea", ""))
    # Usar POC ID + Fecha + Detalle Tarea
    raw = f"{poc_id}|{fecha}|{detalle}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]

def get_mes_actual():
    return datetime.now().strftime("%Y%m")

def formatear_fecha(fecha_valor):
    if pd.isna(fecha_valor):
        return ""
    
    if isinstance(fecha_valor, (datetime, pd.Timestamp)):
        return fecha_valor.strftime("%d/%m/%Y")
    
    fecha_str = str(fecha_valor).strip()
    
    if '/' in fecha_str:
        return fecha_str
    
    if '-' in fecha_str:
        partes = fecha_str.split('-')
        if len(partes) == 3:
            año = partes[0]
            mes = partes[1]
            dia = partes[2].split(' ')[0]
            return f"{dia}/{mes}/{año}"
    
    if len(fecha_str) >= 8 and fecha_str[:8].isdigit():
        año = fecha_str[0:4]
        mes = fecha_str[4:6]
        dia = fecha_str[6:8]
        return f"{dia}/{mes}/{año}"
    
    return fecha_str

def formatear_fecha_revision(fecha_revision):
    """Formatear fecha de revisión a DD/MM/YYYY"""
    if not fecha_revision:
        return ""
    if ' ' in fecha_revision:
        return fecha_revision.split(' ')[0]
    return fecha_revision

def _load_data(path: Path) -> pd.DataFrame:
    global _current_excel, _data_loaded, _current_mes
    
    df = pd.read_excel(path, engine="openpyxl")
    _current_excel = path.name
    _current_mes = get_mes_actual()
    
    if "TaskImageUrl" in df.columns and "Img" not in df.columns:
        df = df.rename(columns={"TaskImageUrl": "Img"})
    
    required = ["Fecha", "Promotor", "POC ID", "Detalle Tarea", "Img", "Completada", "Validada", "Visita Valida", "Supervisor ID"]
    
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Columnas faltantes: {missing}")
    
    df = df.copy()
    
    df["Completada"] = pd.to_numeric(df["Completada"], errors="coerce")
    df["Validada"] = pd.to_numeric(df["Validada"], errors="coerce")
    df["Supervisor ID"] = pd.to_numeric(df["Supervisor ID"], errors="coerce").astype("Int64")
    df["VisitaValidaBool"] = df["Visita Valida"].apply(_is_visita_valida)
    
    filtered = df[
        (df["Completada"] == 1.0) &
        (df["Validada"] == 0.0) &
        (df["VisitaValidaBool"])
    ].copy()
    
    if 'supervisor_id' in session:
        supervisor_id = session['supervisor_id']
        filtered = filtered[filtered["Supervisor ID"] == supervisor_id]
    
    filtered = filtered.reset_index(drop=True)
    filtered["row_id"] = range(1, len(filtered) + 1)
    filtered["task_id"] = filtered.apply(_build_task_id, axis=1)
    
    # === LOG: Mostrar primeros task_id generados ===
    print(f"📊 Task IDs generados: {filtered['task_id'].head().tolist()}")
    print(f"📊 Total tareas: {len(filtered)}")
    # ============================================
    
    _data_loaded = True
    
    return filtered

# =====================================================
# ROUTES
# =====================================================

@app.route("/api/verificar_revisiones")
@login_required
def verificar_revisiones():
    """Ver todas las revisiones guardadas"""
    supervisor_id = session.get('supervisor_id')
    mes = _current_mes or get_mes_actual()
    reviews = db.get_all_reviews(supervisor_id, mes)
    
    return jsonify({
        "total": len(reviews),
        "revisiones": reviews[:10]  # Mostrar las primeras 10
    })

@app.route("/")
def index():
    return render_template("index.html", supervisors=SUPERVISORS)

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json
    supervisor_id = data.get("supervisor_id")
    
    if supervisor_id not in SUPERVISORS:
        return jsonify({"error": "Supervisor no encontrado"}), 404
    
    session['supervisor_id'] = supervisor_id
    session['supervisor_name'] = SUPERVISORS[supervisor_id]
    
    global _cached_df, _data_loaded
    file_path = UPLOAD_DIR / "data.xlsx"
    if file_path.exists():
        try:
            _cached_df = _load_data(file_path)
            _data_loaded = True
        except Exception as e:
            print(f"Error recargando datos: {e}")
    
    return jsonify({
        "ok": True,
        "supervisor_id": supervisor_id,
        "supervisor_name": SUPERVISORS[supervisor_id]
    })

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/has_file")
@login_required
def api_has_file():
    global _data_loaded
    file_exists = (UPLOAD_DIR / "data.xlsx").exists()
    return jsonify({
        "has_file": file_exists and _data_loaded,
        "file_exists": file_exists,
        "data_loaded": _data_loaded,
        "mes_actual": _current_mes
    })

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    global _cached_df, _data_loaded, _current_mes
    
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "El archivo debe ser .xlsx"}), 400
    
    file_path = UPLOAD_DIR / "data.xlsx"
    f.save(str(file_path))
    
    try:
        _cached_df = _load_data(file_path)
        _data_loaded = True
        _current_mes = get_mes_actual()
        
        return jsonify({
            "ok": True,
            "rows": len(_cached_df),
            "message": f"Archivo cargado con {len(_cached_df)} registros",
            "mes": _current_mes
        })
    except Exception as e:
        _cached_df = None
        _data_loaded = False
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks")
@login_required
def api_tasks():
    global _cached_df, _data_loaded
    if not _data_loaded or _cached_df is None or _cached_df.empty:
        file_path = UPLOAD_DIR / "data.xlsx"
        if file_path.exists():
            try:
                _cached_df = _load_data(file_path)
                _data_loaded = True
            except Exception as e:
                return jsonify({"error": f"Error cargando datos: {str(e)}", "no_data": True}), 404
        else:
            return jsonify({"error": "No hay datos cargados", "no_data": True}), 404
    
    supervisor_id = session['supervisor_id']
    start_date = request.args.get("start_date", type=str)
    end_date = request.args.get("end_date", type=str)
    
    result = _cached_df[_cached_df["Supervisor ID"] == supervisor_id].copy()
    
    # Ordenar por fecha (más antiguas primero)
    result = result.sort_values(by="Fecha", ascending=True)
    
    if result.empty:
        return jsonify({"error": f"No hay tareas asignadas para {SUPERVISORS[supervisor_id]}", "no_tasks": True}), 404
    
    # =====================================================
    # FILTROS DE FECHA CORREGIDOS (usando datetime)
    # =====================================================
    if start_date or end_date:
        # Crear columna de fecha como datetime para comparar
        try:
            # Convertir la fecha del Excel a datetime
            result['fecha_dt'] = pd.to_datetime(result['Fecha'], format='%d/%m/%Y', errors='coerce')
            
            if start_date:
                start_dt = datetime.strptime(start_date, "%d/%m/%Y")
                result = result[result['fecha_dt'] >= start_dt]
            
            if end_date:
                end_dt = datetime.strptime(end_date, "%d/%m/%Y")
                result = result[result['fecha_dt'] <= end_dt]
            
            # Eliminar columna temporal
            result = result.drop(columns=['fecha_dt'])
        except Exception as e:
            print(f"Error en filtros de fecha: {e}")
            # Si falla, intentar comparar como string
            if start_date:
                result = result[result["Fecha"].astype(str) >= start_date]
            if end_date:
                result = result[result["Fecha"].astype(str) <= end_date]
    # =====================================================
    
    if result.empty:
        return jsonify({"error": "No hay tareas en el rango de fechas", "no_tasks": True}), 404
    
    task_ids = result["task_id"].tolist()
    pending_tasks = db.get_pending_tasks(supervisor_id, task_ids)
    
    response_rows = []
    for _, row in result.iterrows():
        task_id = row["task_id"]
        is_reviewed = task_id not in pending_tasks
        review = None
        if is_reviewed:
            review = db.get_review_status(task_id, supervisor_id)
        
        raw_poc_id = clean_text(row.get("POC ID"))
        short_poc_id = extract_short_poc_id(raw_poc_id)
        client_info = get_client_info(raw_poc_id)
        
        img_url = clean_text(row.get("Img", ""))
        if not img_url:
            img_url = clean_text(row.get("TaskImageUrl", ""))
        
        response_rows.append({
            "row_id": int(row["row_id"]),
            "task_id": task_id,
            "fecha": formatear_fecha(row.get("Fecha")),
            "promotor": clean_text(row.get("Promotor")),
            "poc_id": short_poc_id,
            "poc_id_completo": raw_poc_id,
            "cliente_id": short_poc_id,
            "razon_social": client_info.get("nombre") if client_info else "SIN DATO",
            "direccion": client_info.get("direccion") if client_info else "SIN DATO",
            "detalle_tarea": clean_text(row.get("Detalle Tarea")),
            "imagen": img_url,
            "revisado": is_reviewed,
            "status": review.get("status") if review else None,
            "observaciones": review.get("observaciones") if review else "",
            "fecha_revision": review.get("fecha_revision") if review else None
        })
    
    return jsonify(response_rows)

# =====================================================
# LAS DEMÁS RUTAS (save_review, delete_review, stats, supervisors, export_reviews, close_month)
# =====================================================

@app.route("/api/save_review", methods=["POST"])
@login_required
def api_save_review():
    data = request.json
    task_id = data.get("task_id")
    status = data.get("status")
    observaciones = data.get("observaciones", "")
    
    if not task_id or not status:
        return jsonify({"error": "Faltan datos"}), 400
    
    if status not in ["objeccion", "invalida", "fraude"]:
        return jsonify({"error": "Status inválido"}), 400
    
    if _cached_df is None or not _data_loaded:
        return jsonify({"error": "No hay datos cargados"}), 404
    
    task_row = _cached_df[_cached_df["task_id"] == task_id]
    if task_row.empty:
        return jsonify({"error": "Tarea no encontrada"}), 404
    
    row_id = int(task_row.iloc[0]["row_id"])
    supervisor_id = session['supervisor_id']
    supervisor_name = session['supervisor_name']
    
    success = db.save_review(
        task_id=task_id,
        row_id=row_id,
        supervisor_id=supervisor_id,
        supervisor_name=supervisor_name,
        status=status,
        observaciones=observaciones,
        excel_file=_current_excel or "data.xlsx",
        mes_revision=_current_mes or get_mes_actual()
    )
    
    if success:
        return jsonify({"ok": True, "message": "Revisión guardada correctamente"})
    else:
        return jsonify({"error": "Error al guardar la revisión"}), 500

@app.route("/api/delete_review/<task_id>", methods=["DELETE"])
@login_required
def api_delete_review(task_id):
    try:
        supervisor_id = session.get('supervisor_id')
        
        if _cached_df is None or not _data_loaded:
            return jsonify({"error": "No hay datos cargados"}), 404
        
        task_row = _cached_df[_cached_df["task_id"] == task_id]
        if task_row.empty:
            return jsonify({"error": "Tarea no encontrada"}), 404
        
        success = db.delete_review(task_id, supervisor_id)
        
        if success:
            return jsonify({"ok": True, "message": "Revisión eliminada correctamente"})
        else:
            return jsonify({"error": "Error al eliminar la revisión"}), 500
            
    except Exception as e:
        print(f"Error al eliminar revisión: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/stats")
@login_required
def api_stats():
    global _data_loaded
    
    supervisor_id = session['supervisor_id']
    mes = _current_mes or get_mes_actual()
    stats = db.get_supervisor_stats(supervisor_id, mes)
    
    if not _data_loaded or _cached_df is None:
        return jsonify({
            "total_revisados": stats["total"],
            "total_pendientes": 0,
            "by_status": stats["by_status"],
            "porcentaje": 0,
            "no_data": True
        })
    
    total_disponibles = len(_cached_df[_cached_df["Supervisor ID"] == supervisor_id])
    task_ids = _cached_df[_cached_df["Supervisor ID"] == supervisor_id]["task_id"].tolist()
    pending_tasks = db.get_pending_tasks(supervisor_id, task_ids)
    total_pendientes = len(pending_tasks)
    
    return jsonify({
        "total_revisados": stats["total"],
        "total_pendientes": total_pendientes,
        "by_status": stats["by_status"],
        "porcentaje": round((stats["total"] / total_disponibles * 100) if total_disponibles > 0 else 0, 1),
        "total_disponibles": total_disponibles,
        "has_data": True,
        "mes": mes
    })

@app.route("/api/supervisors")
def api_supervisors():
    items = [{"id": sid, "name": name} for sid, name in SUPERVISORS.items()]
    return jsonify(items)

@app.route("/api/export_reviews", methods=["GET"])
@login_required
def api_export_reviews():
    try:
        supervisor_id = session.get('supervisor_id')
        supervisor_name = session.get('supervisor_name')
        mes = _current_mes or get_mes_actual()
        
        reviews = db.get_all_reviews(supervisor_id, mes)
        
        if not reviews:
            return jsonify({"error": "No hay revisiones para exportar"}), 404
        
        if _cached_df is None or not _data_loaded:
            export_data = []
            for review in reviews:
                export_data.append({
                    "Fecha Revisión": formatear_fecha_revision(review.get("fecha_revision", "")),
                    "Fecha Tarea": "",
                    "POC ID": "",
                    "Detalle Tarea": "",
                    "URL Imagen": "",
                    "ID Tarea": "",
                    "Razón Social": "",
                    "Direccion": "",
                    "Promotor": "",
                    "Status": review["status"],
                    "Observaciones": review.get("observaciones", "")
                })
            df_export = pd.DataFrame(export_data)
        else:
            task_data = {}
            for _, row in _cached_df.iterrows():
                raw_poc_id = clean_text(row.get("POC ID", ""))
                client_info = get_client_info(raw_poc_id) if raw_poc_id else None
                img_url = clean_text(row.get("Img", ""))
                if not img_url:
                    img_url = clean_text(row.get("TaskImageUrl", ""))
                
                task_data[row["task_id"]] = {
                    "fecha_tarea": formatear_fecha(row.get("Fecha")),
                    "poc_id_completo": raw_poc_id,
                    "detalle_tarea": row.get("Detalle Tarea", ""),
                    "imagen": img_url,
                    "id_tarea": row.get("ID Tarea", ""),  # Campo ID Tarea del Excel
                    "razon_social": client_info.get("nombre") if client_info else "",
                    "direccion": client_info.get("direccion") if client_info else "",
                    "promotor": row.get("Promotor", "")
                }
            
            export_data = []
            for review in reviews:
                task_id = review["task_id"]
                task_info = task_data.get(task_id, {})
                
                export_data.append({
                    "Fecha Revisión": formatear_fecha_revision(review.get("fecha_revision", "")),
                    "Fecha Tarea": task_info.get("fecha_tarea", ""),
                    "POC ID": task_info.get("poc_id_completo", ""),
                    "Detalle Tarea": task_info.get("detalle_tarea", ""),
                    "URL Imagen": task_info.get("imagen", ""),
                    "ID Tarea": task_info.get("id_tarea", ""),
                    "Razón Social": task_info.get("razon_social", ""),
                    "Direccion": task_info.get("direccion", ""),
                    "Promotor": task_info.get("promotor", ""),
                    "Status": review["status"],
                    "Observaciones": review.get("observaciones", "")
                })
            df_export = pd.DataFrame(export_data)
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Revisiones', index=False)
            
            worksheet = writer.sheets['Revisiones']
            
            from openpyxl.styles import PatternFill
            
            # === COLOR AZUL CLARO PARA LAS PRIMERAS 6 COLUMNAS (A a F) ===
            azul_claro = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
            
            for col_idx in range(0, 6):  # A, B, C, D, E, F
                for row in range(1, len(df_export) + 2):
                    cell = worksheet.cell(row=row, column=col_idx + 1)
                    cell.fill = azul_claro
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        filename = f"reporte_revisiones_{supervisor_name}_{mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error al exportar: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/close_month", methods=["POST"])
@login_required
def api_close_month():
    global _cached_df, _data_loaded
    
    try:
        supervisor_id = session.get('supervisor_id')
        supervisor_name = session.get('supervisor_name')
        mes = _current_mes or get_mes_actual()
        
        reviews = db.get_all_reviews(supervisor_id, mes)
        
        if not reviews:
            return jsonify({"error": "No hay revisiones para exportar"}), 404
        
        if _cached_df is None or not _data_loaded:
            export_data = []
            for review in reviews:
                export_data.append({
                    "Fecha Revisión": formatear_fecha_revision(review.get("fecha_revision", "")),
                    "Fecha Tarea": "",
                    "POC ID": "",
                    "Detalle Tarea": "",
                    "URL Imagen": "",
                    "ID Tarea": "",
                    "Razón Social": "",
                    "Direccion": "",
                    "Promotor": "",
                    "Status": review["status"],
                    "Observaciones": review.get("observaciones", "")
                })
            df_export = pd.DataFrame(export_data)
        else:
            task_data = {}
            for _, row in _cached_df.iterrows():
                raw_poc_id = clean_text(row.get("POC ID", ""))
                client_info = get_client_info(raw_poc_id) if raw_poc_id else None
                img_url = clean_text(row.get("Img", ""))
                if not img_url:
                    img_url = clean_text(row.get("TaskImageUrl", ""))
                
                task_data[row["task_id"]] = {
                    "fecha_tarea": formatear_fecha(row.get("Fecha")),
                    "poc_id_completo": raw_poc_id,
                    "detalle_tarea": row.get("Detalle Tarea", ""),
                    "imagen": img_url,
                    "id_tarea": row.get("ID Tarea", ""),
                    "razon_social": client_info.get("nombre") if client_info else "",
                    "direccion": client_info.get("direccion") if client_info else "",
                    "promotor": row.get("Promotor", "")
                }
            
            export_data = []
            for review in reviews:
                task_id = review["task_id"]
                task_info = task_data.get(task_id, {})
                
                export_data.append({
                    "Fecha Revisión": formatear_fecha_revision(review.get("fecha_revision", "")),
                    "Fecha Tarea": task_info.get("fecha_tarea", ""),
                    "POC ID": task_info.get("poc_id_completo", ""),
                    "Detalle Tarea": task_info.get("detalle_tarea", ""),
                    "URL Imagen": task_info.get("imagen", ""),
                    "ID Tarea": task_info.get("id_tarea", ""),
                    "Razón Social": task_info.get("razon_social", ""),
                    "Direccion": task_info.get("direccion", ""),
                    "Promotor": task_info.get("promotor", ""),
                    "Status": review["status"],
                    "Observaciones": review.get("observaciones", "")
                })
            df_export = pd.DataFrame(export_data)
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Revisiones Cierre', index=False)
            
            worksheet = writer.sheets['Revisiones Cierre']
            
            from openpyxl.styles import PatternFill
            
            # === COLOR AZUL CLARO PARA LAS PRIMERAS 6 COLUMNAS (A a F) ===
            azul_claro = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
            
            for col_idx in range(0, 6):  # A, B, C, D, E, F
                for row in range(1, len(df_export) + 2):
                    cell = worksheet.cell(row=row, column=col_idx + 1)
                    cell.fill = azul_claro
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        filename = f"cierre_mes_{supervisor_name}_{mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        db.clear_reviews(supervisor_id, mes)
        
        file_path = UPLOAD_DIR / "data.xlsx"
        if file_path.exists():
            file_path.unlink()
        
        _cached_df = None
        _data_loaded = False
        
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error en cierre de mes: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)